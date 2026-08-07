from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable

from core.report_link_server import ensure_report_server, report_url

HEADERS = [
    "序號", "Band", "BW", "ARFCN", "測試類型", "iPerf方向", "測試秒數",
    "設定開始時間", "設定完成時間", "UE連線狀態", "PHY DL Mbps", "PHY UL Mbps",
    "iPerf平均 Mbps", "iPerf最低 Mbps", "iPerf最高 Mbps", "傳輸量 MB",
    "Download 平均 Mbps", "Download 最低 Mbps", "Download 最高 Mbps", "Download 傳輸量 MB",
    "Upload 平均 Mbps", "Upload 最低 Mbps", "Upload 最高 Mbps", "Upload 傳輸量 MB",
    "結果", "錯誤原因",
]


WIDTHS = {
    "序號": 8,
    "Band": 20,
    "BW": 18,
    "ARFCN": 24,
    "測試類型": 14,
    "iPerf方向": 14,
    "測試秒數": 12,
    "設定開始時間": 21,
    "設定完成時間": 21,
    "UE連線狀態": 14,
    "PHY DL Mbps": 14,
    "PHY UL Mbps": 14,
    "iPerf平均 Mbps": 23,
    "iPerf最低 Mbps": 23,
    "iPerf最高 Mbps": 23,
    "傳輸量 MB": 23,
    "Download 平均 Mbps": 20,
    "Download 最低 Mbps": 20,
    "Download 最高 Mbps": 20,
    "Download 傳輸量 MB": 21,
    "Upload 平均 Mbps": 20,
    "Upload 最低 Mbps": 20,
    "Upload 最高 Mbps": 20,
    "Upload 傳輸量 MB": 21,
    "結果": 10,
    "錯誤原因": 42,
}


def _cell(value: Any) -> Any:
    return "" if value is None else value


def normalize_row(row: dict[str, Any]) -> list[Any]:
    return [_cell(row.get(key)) for key in HEADERS]


def write_reports(output_dir: Path, batch_id: str, rows: Iterable[dict[str, Any]]) -> dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    row_list = list(rows)
    base = output_dir / f"Batch_Result_{batch_id}"
    txt_path = base.with_suffix(".txt")
    xlsx_path = base.with_suffix(".xlsx")

    # TXT and Excel use the exact same headers, ordering and row values.
    with txt_path.open("w", encoding="utf-8-sig", newline="\n") as handle:
        handle.write("\t".join(HEADERS) + "\n")
        for row in row_list:
            values = [
                str(v).replace("\t", " ").replace("\r", " ").replace("\n", " ")
                for v in normalize_row(row)
            ]
            handle.write("\t".join(values) + "\n")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("產生 Excel 需要 openpyxl，請執行 pip install openpyxl。") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Batch Results"
    sheet.freeze_panes = "A2"
    sheet.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, row in enumerate(row_list, start=2):
        for col_index, value in enumerate(normalize_row(row), start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for index, header in enumerate(HEADERS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS.get(header, 16)
    last_column = get_column_letter(len(HEADERS))
    sheet.auto_filter.ref = f"A1:{last_column}{max(1, len(row_list) + 1)}"
    workbook.save(xlsx_path)

    base_url = ensure_report_server(output_dir.parent, output_dir)
    return {
        "xlsx": str(xlsx_path),
        "txt": str(txt_path),
        "xlsx_url": report_url(base_url, xlsx_path),
        "txt_url": report_url(base_url, txt_path),
    }
