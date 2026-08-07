from pathlib import Path
from tempfile import TemporaryDirectory

from core.batch_report import HEADERS, write_reports
from core.batch_worker import _normalize_item, _band_cli


def test_unlimited_1000_rows_and_matching_reports():
    rows = []
    for i in range(1000):
        rows.append({key: (i + 1 if key == "序號" else f"v{i}") for key in HEADERS})
    with TemporaryDirectory() as td:
        reports = write_reports(Path(td), "test", rows)
        txt = Path(reports["txt"])
        xlsx = Path(reports["xlsx"])
        assert txt.exists() and xlsx.exists()
        lines = txt.read_text(encoding="utf-8-sig").splitlines()
        assert len(lines) == 1001
        assert lines[0].split("\t") == HEADERS
        from openpyxl import load_workbook
        ws = load_workbook(xlsx, read_only=True).active
        assert ws.max_row == 1001
        assert [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))] == HEADERS


def test_batch_items_support_lte_nr_endc_and_actions():
    examples = [
        ({"band_config":"1A","bandwidth_config":"20","action":"phy"}, "set_band"),
        ({"band_config":"n78A","bandwidth_config":"100","action":"download"}, "set_nr_band"),
        ({"band_config":"1A_n78A","bandwidth_config":"20_100","action":"tx","duration_sec":30}, "set_nr_band"),
    ]
    for idx, (raw, expected_intent) in enumerate(examples):
        item = _normalize_item(raw, idx)
        intent, cli = _band_cli(item)
        assert intent == expected_intent
        assert cli
